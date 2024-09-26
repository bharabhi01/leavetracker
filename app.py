from flask import Flask, render_template, request, redirect, jsonify
from openpyxl import Workbook, load_workbook
import json
import os

app = Flask(__name__)

# Define the path to the Excel file
EXCEL_FILE = 'employee_leaves.xlsx'

# Create Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Employee Name", "Leave Dates"])  # Adding header row
    workbook.save(EXCEL_FILE)


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        employee_name = request.form['employee_name']
        leave_dates = request.form['leave_dates']

        save_to_excel(employee_name, leave_dates)
        return redirect('/')

    # Load employee data from Excel
    employee_data = load_from_excel()

    employee_names = ["John Doe", "Jane Smith", "Alice Johnson", "Bob Brown"]
    return render_template('index.html', employee_names=employee_names, employee_data=employee_data)


@app.route('/edit/<int:row_id>', methods=['POST'])
def edit(row_id):
    # Get new data from form
    new_employee_name = request.form['employee_name']
    new_leave_dates = request.form['leave_dates']

    # Update the Excel sheet with new values
    update_excel(row_id, new_employee_name, new_leave_dates)

    return redirect('/')


def save_to_excel(employee_name, leave_dates):
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Find the next available row
    next_row = sheet.max_row + 1

    # Write employee details
    sheet.cell(row=next_row, column=1).value = employee_name

    # Convert leave_dates JSON array back to a list
    leave_dates_list = json.loads(leave_dates)

    # Write leave dates in adjacent cells starting from column 2
    for idx, date in enumerate(leave_dates_list):
        sheet.cell(row=next_row, column=2 + idx).value = date

    workbook.save(EXCEL_FILE)


def load_from_excel():
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Read the data from the Excel file into a list
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        employee_name = row[0]
        leave_dates = list(filter(None, row[1:]))  # Filter out empty cells
        data.append({'name': employee_name, 'dates': leave_dates})

    return data


def update_excel(row_id, employee_name, leave_dates):
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Find the row to update based on row_id
    target_row = row_id + 1  # Adjusting for header

    # Update the name and leave dates
    sheet.cell(row=target_row, column=1).value = employee_name

    # Clear old leave dates
    for col in range(2, sheet.max_column + 1):
        sheet.cell(row=target_row, column=col).value = None

    # Convert leave_dates JSON array back to a list
    leave_dates_list = json.loads(leave_dates)

    # Write new leave dates starting from column 2
    for idx, date in enumerate(leave_dates_list):
        sheet.cell(row=target_row, column=2 + idx).value = date

    workbook.save(EXCEL_FILE)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)